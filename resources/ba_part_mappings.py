"""
BrickArchitect Part Mappings Updater

Creates an Excel file with mapping between each Brickarchitect part number
and corresponding one or several Rebrickable part numbers associated.

Step 1: Create list of BA parts
 a) Scrape List of most common lego parts (All Years), i.e. go through
    all the webpages and save in a list only the parts that have a part image.
 b) Scrape all BA category pages and add parts that are NOT in the most
    common list (i.e. parts with no "Overall Rank" value). This captures
    retired and niche parts that don't appear in the most-common pages.

Step 2: Create mapping of RB part nrs for each BA part
 Open the page for a BA part, fetch the RB part nrs associated (up to 8), 
 save in the Excel row for that BA part.
 Loop over all BA parts. The process is resumable.  
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
import time
import random
import os
from pathlib import Path
from datetime import datetime

# ---------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------
MOST_COMMON_URL = "https://brickarchitect.com/parts/most-common-allyears?page={}"
CATEGORIES_URL = "https://brickarchitect.com/parts/"
CATEGORY_SUFFIX = "?retired=1&partstyle=1"
BASE_URL = "https://brickarchitect.com/parts/"
USER_AGENT = "Mozilla/5.0 (compatible; LEGO-mapper/1.0; +https://brickarchitect.com)"
HEADERS = {"User-Agent": USER_AGENT}
TOTAL_PAGES = 39 # current max pages of https://brickarchitect.com/parts/most-common-allyears

# Resumable & checkpoint settings for Phase 2
CHECKPOINT_INTERVAL = 50   # save after this many parts processed in phase 2
DELAY_MIN = 0
DELAY_MAX = 0.1

FIRST_RB_COLUMN = 3
NR_RB_COLUMNS = 80
LAST_RB_COLUMN = FIRST_RB_COLUMN + NR_RB_COLUMNS - 1

# ---------------------------------------------------------------------
# Helper function: Find latest mapping file
# ---------------------------------------------------------------------
def find_latest_mapping_file(resources_dir: Path):
    """
    Find the latest BA vs RB mapping Excel file in the resources directory.
    
    Args:
        resources_dir: Path to the resources directory
    
    Returns:
        Path: Path to the latest mapping file, or None if not found
    """
    import re
    
    # Pattern to match files like "base_part_mapping_2026-01-18.xlsx"
    pattern = re.compile(r"base_part_mapping_(\d{4}-\d{2}-\d{2})\.xlsx")
    
    mapping_files = []
    for file in resources_dir.glob("base_part_mapping_*.xlsx"):
        match = pattern.match(file.name)
        if match:
            date_str = match.group(1)
            mapping_files.append((date_str, file))
    
    if not mapping_files:
        return None
    
    # Sort by date (descending) and return the latest
    mapping_files.sort(reverse=True, key=lambda x: x[0])
    return mapping_files[0][1]


# ---------------------------------------------------------------------
# Helper function: Display available mapping files with part counts
# ---------------------------------------------------------------------
def display_mapping_files_info(resources_dir: Path, count_parts_callback=None):
    """
    Display available mapping files with part counts in Streamlit.
    
    Args:
        resources_dir: Path to the resources directory
        count_parts_callback: Optional callback function(file_path_str) that returns (total_parts, collection_parts)
    
    Returns:
        list: List of tuples (date_str, file_path, is_latest) for all mapping files found
    """
    import re
    import streamlit as st
    
    try:
        # Pattern to match mapping files
        pattern = re.compile(r"base_part_mapping_(\d{4}-\d{2}-\d{2})\.xlsx")
        
        # Find all mapping files
        mapping_files = []
        for file in resources_dir.glob("base_part_mapping_*.xlsx"):
            match = pattern.match(file.name)
            if match:
                date_str = match.group(1)
                mapping_files.append((date_str, file))
        
        # Sort by date (descending)
        mapping_files.sort(reverse=True, key=lambda x: x[0])
        
        # Find the latest (currently used) file
        latest_mapping = find_latest_mapping_file(resources_dir)
        latest_name = latest_mapping.name if latest_mapping else None
        
        # Display each mapping file with part count
        if mapping_files:
            result = []
            for date_str, file in mapping_files:
                is_latest = (file.name == latest_name)
                result.append((date_str, file, is_latest))
                
                try:
                    # Count parts in this mapping file if callback provided
                    if count_parts_callback:
                        total_parts, _ = count_parts_callback(str(file))
                        
                        # Highlight the currently used file
                        if is_latest:
                            st.markdown(f"- **{file.name}** ({total_parts} parts) ✅ *In use*")
                        else:
                            st.markdown(f"- {file.name} ({total_parts} parts)")
                    else:
                        # No callback, just show filename
                        if is_latest:
                            st.markdown(f"- **{file.name}** ✅ *In use*")
                        else:
                            st.markdown(f"- {file.name}")
                except Exception as e:
                    # If counting fails, just show the filename
                    if is_latest:
                        st.markdown(f"- **{file.name}** ✅ *In use*")
                    else:
                        st.markdown(f"- {file.name}")
            
            return result
        else:
            st.markdown("*No mapping files found*")
            return []
    except Exception as e:
        st.warning(f"⚠️ Could not load mapping files information: {e}")
        return []


# ---------------------------------------------------------------------
# Helper function: Extract Rebrickable part numbers
# ---------------------------------------------------------------------
def get_rebrickable_parts(ba_part_number: str, log_callback=None):
    """
    Fetch Rebrickable part numbers for a given BrickArchitect part number.
    
    Args:
        ba_part_number: BrickArchitect part number
        log_callback: Optional callback function(message, status) for logging
    
    Returns:
        list: List of Rebrickable part numbers
    """
    def log(message, status="info"):
        if log_callback:
            log_callback(message, status)
        else:
            print(message)
    
    url = f"{BASE_URL}{ba_part_number}"
    try:
        response = requests.get(url, headers=HEADERS, timeout=10)
        if response.status_code != 200:
            log(f"⚠️ HTTP {response.status_code} for {url}", "warning")
            return []

        soup = BeautifulSoup(response.text, "html.parser")

        # Locate the "Rebrickable:" label
        label_div = soup.find("div", class_="part_detail_label", string=lambda s: s and "Rebrickable" in s)
        if not label_div:
            log(f"⚠️ No 'Rebrickable' section for {ba_part_number}", "warning")
            return []

        # The matching values are in the next sibling with class 'part_detail_value externalparts'
        value_div = label_div.find_next("div", class_="part_detail_value externalparts")
        if not value_div:
            log(f"⚠️ No Rebrickable value block for {ba_part_number}", "warning")
            return []

        # Find all spans that hold the part numbers
        part_spans = value_div.find_all("span", class_="part_num")
        rebrick_parts = [span.get_text(strip=True).lower() for span in part_spans if span.get_text(strip=True)]

        if rebrick_parts:
            log(f"✅ BA {ba_part_number} → RB {rebrick_parts}", "success")
        else:
            log(f"⚠️ No Rebrickable parts found for {ba_part_number}", "warning")

        return rebrick_parts

    except Exception as e:
        log(f"⚠️ Error fetching {ba_part_number}: {e}", "error")
        return []


# ---------------------------------------------------------------------
# Helper function: Fetch all BA parts from BrickArchitect listings
# ---------------------------------------------------------------------
def fetch_ba_parts_from_page(page: int, output_file: Path, existing_parts: set, log_callback=None, stop_flag_callback=None):
    """
    Fetch BA parts from a single page and add to worksheet.
    
    Args:
        page: Page number to fetch
        output_file: Path to save the workbook
        existing_parts: Set of part numbers already in the workbook (updated in-place)
        log_callback: Optional callback function(message, status) for logging
        stop_flag_callback: Optional callback function() that returns True if should stop
    
    Returns:
        int: Number of parts added from this page
    """
    def log(message, status="info"):
        if log_callback:
            log_callback(message, status)
        else:
            print(message)
    
    def should_stop():
        if stop_flag_callback and stop_flag_callback():
            return True
        return False
    
    if should_stop():
        return 0
    
    log(f"🌐 Fetching page {page}/{TOTAL_PAGES}...", "info")
    url = MOST_COMMON_URL.format(page)
    
    try:
        response = requests.get(url, headers=HEADERS, timeout=10)
        if response.status_code != 200:
            log(f"⚠️ Failed to load page {page}", "warning")
            return 0

        soup = BeautifulSoup(response.text, "html.parser")

        # Extract all part rows by their span classes
        parts = soup.find_all("div", class_="tr")  # each part is one row/div
        log(f"   → Found {len(parts)} parts on page {page}", "info")

        # Load workbook for this page
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        parts_added = 0
        for part in parts:
            if should_stop():
                wb.close()
                return parts_added
            
            num_tag = part.find("span", class_="partnum")
            name_tag = part.find("span", class_="partname")
            img_tag = part.find("span", class_="td part_image")
            
            img_url = None
            if img_tag and img_tag.find("img"):
                img_url = img_tag.find("img")["src"]
                if img_url.endswith("noimg.svg"):
                    continue
            
            if num_tag and name_tag:
                partnum = num_tag.get_text(strip=True)
                
                # Skip if already in workbook
                if partnum in existing_parts:
                    continue
                
                partname = name_tag.get_text(strip=True)
                ws.append([partnum, partname])
                existing_parts.add(partnum)
                log(f"   ➕ Added {partnum} - {partname}", "success")
                parts_added += 1

        # Save progress after each page and close workbook
        wb.save(output_file)
        wb.close()
        log(f"💾 Saved progress after page {page}", "info")
        time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))
        
        return parts_added
        
    except Exception as e:
        log(f"❌ Error fetching page {page}: {e}", "error")
        return 0


# ---------------------------------------------------------------------
# Helper function: Fetch all category links from BrickArchitect
# ---------------------------------------------------------------------
def fetch_category_links(log_callback=None):
    """
    Fetch all category page URLs from the BrickArchitect parts index.
    
    Args:
        log_callback: Optional callback function(message, status) for logging
    
    Returns:
        list: List of tuples (category_name, category_url) for each category
    """
    def log(message, status="info"):
        if log_callback:
            log_callback(message, status)
        else:
            print(message)
    
    log("🌐 Fetching category list from BrickArchitect...", "info")
    try:
        response = requests.get(CATEGORIES_URL, headers=HEADERS, timeout=15)
        if response.status_code != 200:
            log(f"⚠️ Failed to load categories page (HTTP {response.status_code})", "warning")
            return []
        
        soup = BeautifulSoup(response.text, "html.parser")
        links = soup.find_all("a", href=True)
        categories = []
        for link in links:
            href = link["href"]
            if "category-" in href:
                name = link.get_text(strip=True)
                categories.append((name, href))
        
        log(f"📂 Found {len(categories)} categories", "info")
        return categories
    except Exception as e:
        log(f"❌ Error fetching categories: {e}", "error")
        return []


# ---------------------------------------------------------------------
# Helper function: Fetch unranked BA parts from a single category page
# ---------------------------------------------------------------------
def fetch_ba_parts_from_category(category_name, category_url, existing_parts, output_file, log_callback=None, stop_flag_callback=None):
    """
    Fetch parts from a category page that have no "Overall Rank" value
    and are not already in the workbook.
    
    Parts with no Overall Rank are those not appearing in the most-common
    pages (typically retired or niche parts).
    
    Args:
        category_name: Display name of the category
        category_url: Base URL of the category page
        existing_parts: Set of part numbers already in the workbook
        output_file: Path to the output Excel file
        log_callback: Optional callback function(message, status) for logging
        stop_flag_callback: Optional callback function() that returns True if should stop
    
    Returns:
        int: Number of new parts added from this category
    """
    def log(message, status="info"):
        if log_callback:
            log_callback(message, status)
        else:
            print(message)
    
    def should_stop():
        if stop_flag_callback and stop_flag_callback():
            return True
        return False
    
    if should_stop():
        return 0
    
    # Append suffix to show all parts (retired + current) in table view
    url = category_url + CATEGORY_SUFFIX
    log(f"🌐 Fetching category '{category_name}': {url}", "info")
    
    try:
        response = requests.get(url, headers=HEADERS, timeout=15)
        if response.status_code != 200:
            log(f"⚠️ Failed to load category '{category_name}' (HTTP {response.status_code})", "warning")
            return 0
        
        soup = BeautifulSoup(response.text, "html.parser")
        rows = soup.find_all("div", class_="tr")
        
        if len(rows) <= 1:
            log(f"   → No part rows found in category '{category_name}'", "info")
            return 0
        
        # Load workbook to append new parts
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        
        parts_added = 0
        for row in rows[1:]:  # skip header row
            if should_stop():
                break
            
            # Check if "Overall Rank" column is empty
            rank_span = row.find("span", class_="weighted_rank")
            rank_text = rank_span.get_text(strip=True) if rank_span else ""
            if rank_text:
                # Part has a rank → already in most-common pages, skip
                continue
            
            # Extract part details
            num_tag = row.find("span", class_="partnum")
            name_tag = row.find("span", class_="partname")
            img_tag = row.find("span", class_="td part_image")
            
            if not num_tag:
                continue
            
            partnum = num_tag.get_text(strip=True)
            
            # Skip if already in workbook
            if partnum in existing_parts:
                continue
            
            # Must have an actual image (not noimg.svg)
            img_url = None
            if img_tag and img_tag.find("img"):
                img_url = img_tag.find("img")["src"]
                if img_url.endswith("noimg.svg"):
                    continue
            else:
                continue
            
            partname = name_tag.get_text(strip=True) if name_tag else ""
            
            ws.append([partnum, partname])
            existing_parts.add(partnum)
            parts_added += 1
            log(f"   ➕ Added {partnum} - {partname} (from {category_name})", "success")
        
        wb.save(output_file)
        wb.close()
        
        if parts_added > 0:
            log(f"💾 Saved {parts_added} new parts from category '{category_name}'", "info")
        else:
            log(f"   → No new unranked parts in category '{category_name}'", "info")
        
        time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))
        return parts_added
        
    except Exception as e:
        log(f"❌ Error fetching category '{category_name}': {e}", "error")
        return 0


# ---------------------------------------------------------------------
# Phase 1: Fetch all BA parts (resumable)
# ---------------------------------------------------------------------
def fetch_all_ba_parts(output_file: Path, start_page=1, log_callback=None, stop_flag_callback=None, stats_callback=None):
    """
    Fetch all BA parts from BrickArchitect listings (Phase 1).
    
    First scrapes the most-common pages (page 1 to TOTAL_PAGES), then
    scrapes all category pages to pick up unranked parts (retired/niche)
    that don't appear in the most-common list.
    
    Progress is tracked in a JSON file alongside the Excel file so that
    the process can resume correctly even if interrupted by Streamlit reruns.
    
    Args:
        output_file: Path to the output Excel file
        start_page: Page to start from (for resuming, used only on first run)
        log_callback: Optional callback function(message, status) for logging
        stop_flag_callback: Optional callback function() that returns True if should stop
        stats_callback: Optional callback function(stats) to update stats in real-time
    
    Returns:
        dict: Statistics about the fetch process
    """
    import json
    
    stats = {"phase": 1, "pages_processed": 0, "last_page_completed": 0, "parts_added": 0, "categories_processed": 0, "category_parts_added": 0, "stopped": False}
    
    def log(message, status="info"):
        if log_callback:
            log_callback(message, status)
        else:
            print(message)
    
    def update_stats():
        if stats_callback:
            stats_callback(stats.copy())
    
    # Progress file for resumable tracking
    progress_file = output_file.with_suffix(".progress.json")
    
    def load_progress():
        """Load progress from file, returns dict with last_page and categories_done."""
        if progress_file.exists():
            try:
                with open(progress_file, "r") as f:
                    return json.load(f)
            except Exception:
                pass
        return {"last_page_completed": 0, "categories_done": False}
    
    def save_progress(last_page, categories_done=False):
        """Save progress to file."""
        try:
            with open(progress_file, "w") as f:
                json.dump({"last_page_completed": last_page, "categories_done": categories_done}, f)
        except Exception:
            pass
    
    def clear_progress():
        """Remove progress file when Phase 1 is fully complete."""
        try:
            if progress_file.exists():
                progress_file.unlink()
        except Exception:
            pass
    
    # Create or open workbook
    if output_file.exists():
        log(f"🔄 Resuming from existing workbook: {output_file.name}", "info")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        # Create header row with all columns
        header_row = ["BA partnum", "BA partname"]
        # Add RB part columns dynamically based on NR_RB_COLUMNS
        header_row.extend([f"RB part_{i+1}" for i in range(NR_RB_COLUMNS)])
        ws.append(header_row)
        wb.save(output_file)
        wb.close()
        log(f"🆕 Created new workbook: {output_file.name}", "info")
    
    # Build set of existing part numbers to avoid duplicates on resume
    existing_parts = set()
    try:
        wb = openpyxl.load_workbook(output_file, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                existing_parts.add(str(row[0]).strip())
        wb.close()
    except Exception as e:
        log(f"⚠️ Could not read existing parts: {e}", "warning")
    
    if existing_parts:
        log(f"📊 {len(existing_parts)} parts already in workbook", "info")
    
    # Determine resume point from progress file
    progress = load_progress()
    resume_page = progress["last_page_completed"] + 1
    categories_already_done = progress["categories_done"]
    log(f"📄 Progress file: last_page_completed={progress['last_page_completed']}, categories_done={categories_already_done}", "info")
    
    # If progress file says we're past start_page, use the progress file
    if resume_page > start_page:
        start_page = resume_page
        log(f"📌 Resuming most-common pages from page {start_page}", "info")
    
    # --- Step 1a: Fetch most-common pages ---
    if start_page <= TOTAL_PAGES:
        log("📋 Phase 1a: Fetching most-common parts pages...", "info")
        for page in range(start_page, TOTAL_PAGES + 1):
            parts_added = fetch_ba_parts_from_page(page, output_file, existing_parts, log_callback, stop_flag_callback)
            stats["pages_processed"] += 1
            stats["last_page_completed"] = page
            stats["parts_added"] += parts_added
            save_progress(page, categories_done=False)
            update_stats()
            
            if stop_flag_callback and stop_flag_callback():
                log(f"⏹️ Phase 1 stopped by user after page {page}", "warning")
                stats["stopped"] = True
                update_stats()
                break
            
            log(f"✅ Page {page} done, continuing to next page...", "info")
        
        if stats["stopped"]:
            return stats
    else:
        log("📋 Phase 1a: Most-common pages already completed, skipping.", "info")
    
    # --- Step 1b: Fetch unranked parts from category pages ---
    if not categories_already_done:
        log("📋 Phase 1b: Fetching unranked parts from category pages...", "info")
        log(f"📊 {len(existing_parts)} parts in workbook before category scan", "info")
        
        # Fetch category links
        categories = fetch_category_links(log_callback)
        
        for cat_name, cat_url in categories:
            if stop_flag_callback and stop_flag_callback():
                log("⏹️ Phase 1 stopped by user during category scraping", "warning")
                stats["stopped"] = True
                update_stats()
                break
            
            cat_parts = fetch_ba_parts_from_category(
                cat_name, cat_url, existing_parts, output_file,
                log_callback, stop_flag_callback
            )
            stats["categories_processed"] += 1
            stats["category_parts_added"] += cat_parts
            stats["parts_added"] += cat_parts
            update_stats()
        
        if not stats["stopped"]:
            save_progress(TOTAL_PAGES, categories_done=True)
    else:
        log("📋 Phase 1b: Category pages already completed, skipping.", "info")
    
    if not stats["stopped"]:
        # All done — clean up progress file
        clear_progress()
        log(
            f"✅ Phase 1 completed! "
            f"Processed {stats['pages_processed']} most-common pages + {stats['categories_processed']} categories, "
            f"added {stats['parts_added']} parts total ({stats['category_parts_added']} from categories)",
            "success"
        )
    
    return stats


# ---------------------------------------------------------------------
# Phase 2: Resumable Rebrickable fetching
# ---------------------------------------------------------------------
def fetch_rebrickable_mappings(output_file: Path, checkpoint_interval=CHECKPOINT_INTERVAL, log_callback=None, stop_flag_callback=None, stats_callback=None):
    """
    Fetch Rebrickable mappings for BA parts (Phase 2).
    
    Args:
        output_file: Path to the Excel file
        checkpoint_interval: Save after this many parts processed
        log_callback: Optional callback function(message, status) for logging
        stop_flag_callback: Optional callback function() that returns True if should stop
        stats_callback: Optional callback function(stats) to update stats in real-time
    
    Returns:
        dict: Statistics about the fetch process
    """
    stats = {"phase": 2, "total": 0, "processed": 0, "stopped": False}
    
    def log(message, status="info"):
        if log_callback:
            log_callback(message, status)
        else:
            print(message)
    
    def should_stop():
        if stop_flag_callback and stop_flag_callback():
            return True
        return False
    
    def update_stats():
        if stats_callback:
            stats_callback(stats.copy())
    
    # Load workbook
    try:
        wb = openpyxl.load_workbook(output_file, read_only=False, keep_vba=False)
        ws = wb.active
    except Exception as e:
        log(f"❌ Error loading Excel file: {e}", "error")
        raise
    
    # Determine first row that needs processing
    start_row = None
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        # if first RB column (col 5) is empty, that's a row to process
        if not (ws.cell(row=r, column=FIRST_RB_COLUMN).value):
            start_row = r
            break

    if start_row is None:
        log("✅ Phase 2: All rows already have Rebrickable data (nothing to do).", "success")
        return stats

    # Count how many need processing for progress reporting
    rows_to_process = [r for r in range(start_row, max_row + 1) if ws.cell(row=r, column=1).value]
    stats["total"] = len(rows_to_process)
    log(f"🔍 Phase 2: Resuming Rebrickable mapping from row {start_row}. {stats['total']} parts to process.", "info")
    update_stats()

    try:
        for idx, r in enumerate(rows_to_process, start=1):
            if should_stop():
                log("⏹️ Phase 2 stopped by user", "warning")
                stats["stopped"] = True
                wb.save(output_file)
                wb.close()
                update_stats()
                break
            
            ba_part = ws.cell(row=r, column=1).value
            if not ba_part:
                log(f"   ↩️ Row {r} empty BA part, skipping", "info")
                continue

            # Double-check skip if some RB cells are already present
            existing_rb = any(ws.cell(row=r, column=c).value for c in range(FIRST_RB_COLUMN, LAST_RB_COLUMN + 1))
            if existing_rb:
                log(f"   ⏩ Row {r} ({ba_part}) already has RB data, skipping", "info")
                stats["processed"] += 1
                update_stats()
                continue

            log(f"[{idx}/{stats['total']}] 🔎 Fetching RB for row {r}: {ba_part}", "info")
            rb_parts = get_rebrickable_parts(ba_part, log_callback)

            if rb_parts:
                for i, rb in enumerate(rb_parts[:NR_RB_COLUMNS]):
                    ws.cell(row=r, column=FIRST_RB_COLUMN + i, value=rb)
                # Save immediately after writing data
                wb.save(output_file)
            else:
                # Mark explicitly as not found to avoid refetching each run
                ws.cell(row=r, column=FIRST_RB_COLUMN, value="N/A")
                # Save immediately after writing data
                wb.save(output_file)

            stats["processed"] += 1
            update_stats()

            # Checkpoint save message (file already saved above)
            if stats["processed"] % checkpoint_interval == 0:
                log(f"💾 Checkpoint: {stats['processed']} parts processed (row {r})", "info")

            # Delay between part requests
            time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

    except Exception as e:
        log(f"❌ Unexpected error during Phase 2: {e}. Saving progress.", "error")
        wb.save(output_file)
        wb.close()
        raise

    # Final save and close
    wb.save(output_file)
    wb.close()
    if not stats["stopped"]:
        log(f"✅ Phase 2 completed. Processed {stats['processed']} items. Results saved to {output_file.name}", "success")
    else:
        log(f"⏹️ Phase 2 stopped. Progress saved: {stats['processed']} items processed.", "warning")
    
    return stats


# ---------------------------------------------------------------------
# Main update function (for integration with app.py)
# ---------------------------------------------------------------------
def update_ba_mappings(output_file: Path, log_callback=None, stop_flag_callback=None, stats_callback=None):
    """
    Update BA part mappings (both phases).
    
    Args:
        output_file: Path to the output Excel file
        log_callback: Optional callback function(message, status) for logging
        stop_flag_callback: Optional callback function() that returns True if should stop
        stats_callback: Optional callback function(stats) to update stats in real-time
    
    Returns:
        dict: Combined statistics from both phases
    """
    def log(message, status="info"):
        if log_callback:
            log_callback(message, status)
        else:
            print(message)
    
    log("🚀 Starting BA mappings update...", "info")
    
    # Phase 1: Fetch BA parts
    log("📋 Phase 1: Fetching BA parts from BrickArchitect...", "info")
    phase1_stats = fetch_all_ba_parts(output_file, start_page=1, log_callback=log_callback, stop_flag_callback=stop_flag_callback, stats_callback=stats_callback)
    
    if phase1_stats["stopped"]:
        log("⏹️ Update stopped during Phase 1", "warning")
        return phase1_stats
    
    # Phase 2: Fetch Rebrickable mappings
    log("🔗 Phase 2: Fetching Rebrickable mappings...", "info")
    phase2_stats = fetch_rebrickable_mappings(output_file, checkpoint_interval=CHECKPOINT_INTERVAL, log_callback=log_callback, stop_flag_callback=stop_flag_callback, stats_callback=stats_callback)
    
    if not phase2_stats["stopped"]:
        log(f"🎉 Update complete! Results saved to {output_file.name}", "success")
    
    return {
        "phase1": phase1_stats,
        "phase2": phase2_stats,
        "stopped": phase1_stats["stopped"] or phase2_stats["stopped"]
    }


# ---------------------------------------------------------------------
# Command-line execution
# ---------------------------------------------------------------------
if __name__ == "__main__":
    import sys
    
    SCRIPT_DIR = Path(__file__).parent
    timestamp = datetime.now().strftime("%Y-%m-%d")
    OUTPUT_FILE = SCRIPT_DIR / f"base_part_mapping_{timestamp}.xlsx"
    
    # Check command line arguments
    if len(sys.argv) > 1:
        if sys.argv[1] == "phase1":
            print(f"📝 Running Phase 1 only: Fetching BA parts")
            print(f"📝 Output file: {OUTPUT_FILE}")
            fetch_all_ba_parts(OUTPUT_FILE)
        elif sys.argv[1] == "phase2":
            print(f"📝 Running Phase 2 only: Fetching Rebrickable mappings")
            latest_file = find_latest_mapping_file(SCRIPT_DIR)
            if latest_file:
                print(f"📝 Updating file: {latest_file}")
                fetch_rebrickable_mappings(latest_file)
            else:
                print("❌ No mapping file found. Please run phase1 first.")
        else:
            print("Usage: python ba_part_mappings.py [phase1|phase2]")
            print("  phase1: Fetch BA parts only")
            print("  phase2: Fetch Rebrickable mappings only")
            print("  (no argument): Run both phases")
    else:
        # Run both phases
        print(f"📝 Output file: {OUTPUT_FILE}")
        update_ba_mappings(OUTPUT_FILE)
