"""
BrickArchitect Label Downloader

This module downloads label files (.lbx) from BrickArchitect based on the
part mapping Excel file. Labels are cached locally to avoid re-downloading.
"""

import os
import requests
from pathlib import Path
import openpyxl


def download_ba_labels(mapping_path: Path, cache_labels_dir: Path, timeout: int = 10, progress_callback=None, stop_flag_callback=None, stats_callback=None, filter_mode: str = "all", collection_parts: set = None):
    """
    Download BrickArchitect label files based on the mapping Excel file.
    
    Args:
        mapping_path: Path to the Excel file containing BA part mappings
        cache_labels_dir: Directory to cache downloaded label files
        timeout: Request timeout in seconds
        progress_callback: Optional callback function(message, status) for progress updates
            status can be: 'info', 'success', 'warning', 'error'
        stop_flag_callback: Optional callback function() that returns True if download should stop
        stats_callback: Optional callback function(stats) to update stats in real-time
        filter_mode: "all" to download all parts, "collection" to download only parts in collection
        collection_parts: Set of RB part numbers from user's collection (required if filter_mode="collection")
    
    Returns:
        dict: Statistics about the download process
            - total: Total number of label URLs found
            - skipped: Number of files already cached
            - downloaded: Number of files successfully downloaded
            - failed: Number of failed downloads
            - stopped: Whether the download was stopped by user
            - filtered: Number of parts filtered out (when filter_mode="collection")
    """
    stats = {"total": 0, "skipped": 0, "downloaded": 0, "failed": 0, "stopped": False, "filtered": 0}
    
    def log(message, status="info"):
        if progress_callback:
            progress_callback(message, status)
        else:
            print(message)
    
    def should_stop():
        """Check if user requested to stop"""
        if stop_flag_callback and stop_flag_callback():
            return True
        return False
    
    def update_stats():
        """Update stats via callback if provided"""
        if stats_callback:
            stats_callback(stats.copy())
    
    # Setup
    os.makedirs(cache_labels_dir, exist_ok=True)
    
    # Load workbook
    try:
        wb = openpyxl.load_workbook(mapping_path)
        ws = wb.active
    except Exception as e:
        log(f"❌ Error loading Excel file: {e}", "error")
        raise
    
    # Identify columns
    header_row = [cell.value for cell in ws[1]]
    try:
        partnum_col = header_row.index("BA partnum") + 1
    except ValueError as e:
        log("❌ Could not find required column 'BA partnum'", "error")
        raise ValueError("Could not find required column 'BA partnum'") from e
    
    # Find all RB part columns (RB part_1, RB part_2, etc.)
    rb_part_cols = []
    for idx, col_name in enumerate(header_row):
        if col_name and col_name.startswith("RB part_"):
            rb_part_cols.append(idx + 1)
    
    # Validate filter mode
    if filter_mode == "collection":
        if collection_parts is None:
            log("❌ filter_mode='collection' requires collection_parts parameter", "error")
            raise ValueError("filter_mode='collection' requires collection_parts parameter")
        if not rb_part_cols:
            log("❌ filter_mode='collection' requires 'RB part_' columns in mapping file", "error")
            raise ValueError("filter_mode='collection' requires 'RB part_' columns in mapping file")
    
    # Download each label file
    for row in ws.iter_rows(min_row=2):
        # Check if user wants to stop
        if should_stop():
            log("⏹️ Download stopped by user", "warning")
            stats["stopped"] = True
            update_stats()
            break
        
        partnum = row[partnum_col - 1].value
        
        if not partnum:
            continue
        
        partnum = str(partnum).strip()
        
        # Filter by collection if requested
        if filter_mode == "collection" and rb_part_cols:
            # Check if any RB part matches collection
            found_in_collection = False
            for col_idx in rb_part_cols:
                rb_partnum = str(row[col_idx - 1].value).strip() if row[col_idx - 1].value else None
                if rb_partnum and rb_partnum in collection_parts:
                    found_in_collection = True
                    break
            
            if not found_in_collection:
                stats["filtered"] += 1
                update_stats()
                continue
        
        # Construct label URL from part number (fixed pattern)
        label_url = f"https://brickarchitect.com/label/{partnum}.lbx"
        filename = f"{partnum}.lbx"
        
        stats["total"] += 1
        update_stats()
        
        save_path = os.path.join(cache_labels_dir, filename)
        
        # Skip if already downloaded
        if os.path.exists(save_path):
            log(f"✅ Skipping (already exists): {filename}", "info")
            stats["skipped"] += 1
            update_stats()
            continue
        
        # Check again before downloading (user might have clicked stop)
        if should_stop():
            log("⏹️ Download stopped by user", "warning")
            stats["stopped"] = True
            update_stats()
            break
        
        # Download file
        try:
            log(f"⬇️ Downloading {filename}...", "info")
            response = requests.get(label_url, timeout=timeout)
            if response.status_code == 200 and response.content:
                with open(save_path, "wb") as f:
                    f.write(response.content)
                log(f"✅ Saved: {filename}", "success")
                stats["downloaded"] += 1
                update_stats()
            else:
                log(f"⚠️ Failed to download {filename} (status: {response.status_code})", "warning")
                stats["failed"] += 1
                update_stats()
        except Exception as e:
            log(f"❌ Error downloading {filename}: {e}", "error")
            stats["failed"] += 1
            update_stats()
    
    # Summary
    if stats["stopped"]:
        log(f"⏹️ Download stopped!", "warning")
    else:
        log(f"🎉 Download complete!", "success")
    
    return stats


# Command-line execution
if __name__ == "__main__":
    from ba_part_mappings import find_latest_mapping_file
    
    SCRIPT_DIR = Path(__file__).parent
    INPUT_FILE = find_latest_mapping_file(SCRIPT_DIR)
    
    if not INPUT_FILE:
        print("❌ No mapping file found in resources directory")
        exit(1)
    
    print(f"📂 Using mapping file: {INPUT_FILE.name}")
    GLOBAL_CACHE_DIR = SCRIPT_DIR.parent / "cache"
    CACHE_LABELS_DIR = GLOBAL_CACHE_DIR / "labels"
    
    download_ba_labels(INPUT_FILE, CACHE_LABELS_DIR)
