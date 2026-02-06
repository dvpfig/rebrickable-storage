# ---------------------------------------------------------------------
# 
#  Creates a Excel file with mapping between each Brickarchitect part number
#  and corresponding one or several Rebrickable part numbers associated.
#  
#  Step 1: Create list of BA parts
#   Scrape List of most common lego parts (All Years), i.e. go through
#   all the webpages and save in a list only the parts that have a part image. 
#
#  Step 2: Create mapping of RB part nrs for each BA part
#   Open the page for a BA part, fetch the RB part nrs associated (up to 8), 
#   save in the Excel row for that BA part.
#   Loop over all BA parts. The process is resumable.  
#
# ---------------------------------------------------------------------

import requests
from bs4 import BeautifulSoup
import openpyxl
import time
import random
import os

# ---------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------
#OUTPUT_FILE = "part number - BA vs RB - 2025-11-11.xlsx"
OUTPUT_FILE = "part number - BA vs RB - 2026-01-18.xlsx"
MOST_COMMON_URL = "https://brickarchitect.com/parts/most-common-allyears?page={}"
BASE_URL = "https://brickarchitect.com/parts/"
USER_AGENT = "Mozilla/5.0 (compatible; LEGO-mapper/1.0; +https://brickarchitect.com)"
HEADERS = {"User-Agent": USER_AGENT}
TOTAL_PAGES = 39

# Resumable & checkpoint settings for Phase 2
CHECKPOINT_INTERVAL = 50   # save after this many parts processed in phase 2
DELAY_MIN = 0
DELAY_MAX = 0.1

# ---------------------------------------------------------------------
# Create or open workbook
# ---------------------------------------------------------------------
if os.path.exists(OUTPUT_FILE):
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb.active
    print(f"🔄 Resuming from existing workbook: {OUTPUT_FILE}")
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["BA partnum", "BA partname", "BA image URL", "BA label URL"])
    ws.append(["RB part_1", "RB part_2", "RB part_3", "RB part_4", "RB part_5", "RB part_6", "RB part_7", "RB part_8", "RB part_9", "RB part_10"])
    wb.save(OUTPUT_FILE)
    print(f"🆕 Created new workbook: {OUTPUT_FILE}")

FIRST_RB_COLUMN = 5
NR_RB_COLUMNS = 10
LAST_RB_COLUMN = 14

# ---------------------------------------------------------------------
# Helper function: Extract Rebrickable part numbers
# ---------------------------------------------------------------------
def get_rebrickable_parts(ba_part_number: str):
    url = f"{BASE_URL}{ba_part_number}"
    try:
        response = requests.get(url, headers=HEADERS, timeout=10)
        if response.status_code != 200:
            print(f"⚠️ HTTP {response.status_code} for {url}")
            return []

        soup = BeautifulSoup(response.text, "html.parser")

        # Locate the "Rebrickable:" label
        label_div = soup.find("div", class_="part_detail_label", string=lambda s: s and "Rebrickable" in s)
        if not label_div:
            print(f"⚠️ No 'Rebrickable' section for {ba_part_number}")
            return []

        # The matching values are in the next sibling with class 'part_detail_value externalparts'
        value_div = label_div.find_next("div", class_="part_detail_value externalparts")
        if not value_div:
            print(f"⚠️ No Rebrickable value block for {ba_part_number}")
            return []

        # Find all spans that hold the part numbers
        part_spans = value_div.find_all("span", class_="part_num")
        rebrick_parts = [span.get_text(strip=True).lower() for span in part_spans if span.get_text(strip=True)]

        if rebrick_parts:
            print(f"✅ BA {ba_part_number} → RB {rebrick_parts}")
        else:
            print(f"⚠️ No Rebrickable parts found for {ba_part_number}")

        return rebrick_parts

    except Exception as e:
        print(f"⚠️ Error fetching {ba_part_number}: {e}")
        return []


# ---------------------------------------------------------------------
# Helper function: Fetch all BA parts from BrickArchitect listings
# ---------------------------------------------------------------------
def get_all_ba_parts(start_page=1):
    for page in range(start_page, TOTAL_PAGES + 1):
        print(f"\n🌐 Fetching page {page}/{TOTAL_PAGES}...")
        url = MOST_COMMON_URL.format(page)
        response = requests.get(url, headers=HEADERS)
        if response.status_code != 200:
            print(f"⚠️ Failed to load {url}")
            continue

        soup = BeautifulSoup(response.text, "html.parser")

        # Extract all part rows by their span classes
        parts = soup.find_all("div", class_="tr")  # each part is one row/div
        print(f"   → Found {len(parts)} parts on this page")

        for part in parts:
            num_tag = part.find("span", class_="partnum")
            name_tag = part.find("span", class_="partname")
            img_tag = part.find("span", class_="td part_image")
            
            img_url = None
            if img_tag and img_tag.find("img"):
                img_url = img_tag.find("img")["src"]
                if img_url.endswith("noimg.svg"):
                    continue
            
            if num_tag and name_tag:
                partnum = num_tag.get_text(strip=True)
                partname = name_tag.get_text(strip=True)
                label_url = f"https://brickarchitect.com/label/{partnum}.lbx"
                ws.append([partnum, partname, img_url, label_url])
                print(f"   ➕ Added {partnum} - {partname}")

        # Save progress after each page
        wb.save(OUTPUT_FILE)
        print(f"💾 Saved progress after page {page}")
        time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

# ---------------------------------------------------------------------
# Phase 2: Resumable Rebrickable fetching
# ---------------------------------------------------------------------
def resume_and_fill_rebrickable(ws, wb, checkpoint_interval=CHECKPOINT_INTERVAL):
    # determine first row that needs processing:
    start_row = None
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        # if first RB column (col 5) is empty, that's a row to process
        if not (ws.cell(row=r, column=FIRST_RB_COLUMN).value):
            start_row = r
            break

    if start_row is None:
        print("✅ Phase 2: All rows already have Rebrickable data (nothing to do).")
        return

    # count how many need processing for progress reporting
    rows_to_process = [r for r in range(start_row, max_row + 1) if ws.cell(row=r, column=1).value]
    total = len(rows_to_process)
    print(f"\n🔍 Phase 2: Resuming Rebrickable mapping from row {start_row}. {total} parts to process.")

    processed = 0
    try:
        for idx, r in enumerate(rows_to_process, start=1):
            ba_part = ws.cell(row=r, column=1).value
            if not ba_part:
                print(f"   ↩️ Row {r} empty BA part, skipping")
                continue

            # double-check skip if some RB cells are already present
            existing_rb = any(ws.cell(row=r, column=c).value for c in range(FIRST_RB_COLUMN, LAST_RB_COLUMN))
            if existing_rb:
                print(f"   ⏩ Row {r} ({ba_part}) already has RB data, skipping")
                processed += 1
                continue

            print(f"[{idx}/{total}] 🔎 Fetching RB for row {r}: {ba_part}")
            rb_parts = get_rebrickable_parts(ba_part)

            if rb_parts:
                for i, rb in enumerate(rb_parts[:NR_RB_COLUMNS]):
                    ws.cell(row=r, column=FIRST_RB_COLUMN + i, value=rb)
            else:
                # mark explicitly as not found to avoid refetching each run
                ws.cell(row=r, column=FIRST_RB_COLUMN, value="N/A")

            processed += 1

            # checkpoint save
            if processed % checkpoint_interval == 0:
                wb.save(OUTPUT_FILE)
                print(f"💾 Checkpoint saved after {processed} processed (row {r})")

            # delay between part requests
            time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

    except KeyboardInterrupt:
        print("\n⏸️ Interrupted by user — saving progress and exiting.")
        wb.save(OUTPUT_FILE)
        raise

    except Exception as e:
        print(f"\n⚠️ Unexpected error during Phase 2: {e}. Saving progress.")
        wb.save(OUTPUT_FILE)
        raise

    # final save
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Phase 2 completed. Processed {processed} items. Results saved to {OUTPUT_FILE}")

# ---------------------------------------------------------------------
# Main logic
# ---------------------------------------------------------------------
# Check how many rows already exist (for resuming BA scraping)
existing_rows = ws.max_row
if existing_rows > 1:
    # Each page has around 250 parts — estimate start page
    print(f"Using BA scraping workbook with total rows: {existing_rows}")
else: 
    # Phase 1: BA scraping 
    get_all_ba_parts(1)

# Phase 2: Rebrickable fetching (resumable & checkpointed)
print("\n🔍 Starting Rebrickable mapping phase (resumable)...")
resume_and_fill_rebrickable(ws, wb, checkpoint_interval=CHECKPOINT_INTERVAL)

print(f"\n✅ Done! Results saved to {OUTPUT_FILE}")
