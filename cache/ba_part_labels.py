import os
import requests
from pathlib import Path
import openpyxl
from urllib.parse import urlparse

# ---------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------
RESOURCES_DIR = Path("resources")
INPUT_FILE = RESOURCES_DIR / "part number - BA vs RB - 2025-11-11.xlsx"

GLOBAL_CACHE_DIR = Path("cache")
CACHE_LABELS_DIR = GLOBAL_CACHE_DIR / "labels"
TIMEOUT = 10

# ---------------------------------------------------------------------
# Setup
# ---------------------------------------------------------------------
os.makedirs(CACHE_LABELS_DIR, exist_ok=True)

wb = openpyxl.load_workbook(INPUT_FILE)
ws = wb.active

# Identify columns
header_row = [cell.value for cell in ws[1]]
try:
    partnum_col = header_row.index("BA partnum") + 1
    labelurl_col = header_row.index("BA label URL") + 1
except ValueError as e:
    raise ValueError("Could not find required columns 'BA partnum' or 'BA label URL'") from e

# ---------------------------------------------------------------------
# Download each label file
# ---------------------------------------------------------------------
for row in ws.iter_rows(min_row=2):
    partnum = row[partnum_col - 1].value
    label_url = row[labelurl_col - 1].value

    if not label_url or "No label available" in str(label_url):
        continue

    # Extract filename from URL
    parsed = urlparse(label_url)
    filename = os.path.basename(parsed.path)

    if not filename.lower().endswith(".lbx"):
        continue

    save_path = os.path.join(CACHE_LABELS_DIR, filename)

    # Skip if already downloaded
    if os.path.exists(save_path):
        print(f"✅ Skipping (already exists): {filename}")
        continue

    # Download file
    try:
        print(f"⬇️  Downloading {label_url} ...")
        response = requests.get(label_url, timeout=TIMEOUT)
        if response.status_code == 200 and response.content:
            with open(save_path, "wb") as f:
                f.write(response.content)
            print(f"✅ Saved: {save_path}")
        else:
            print(f"⚠️  Failed to download {label_url} (status: {response.status_code})")
    except Exception as e:
        print(f"❌ Error downloading {label_url}: {e}")

print("\n🎉 All downloads complete.")
