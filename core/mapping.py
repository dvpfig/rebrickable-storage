# core/mapping.py
import streamlit as st
import pandas as pd
from io import BytesIO
from pathlib import Path
from streamlit import cache_data

from core.custom_mapping import (
    load_custom_mapping_csv,
    build_custom_mapping_dict,
    apply_custom_mapping
)







class EnhancedMapping(dict):
    """
    Enhanced mapping dictionary that applies custom rules as fallback.

    Priority:
    1. Check explicit Excel mappings (base_mapping) for original RB part
    2. Check custom mapping (exact and wildcard patterns)
    3. Return original key if no mapping found
    
    Note: Parts not found in either Excel or custom mapping will return the original
    RB part number, allowing the application to attempt retrieval from Rebrickable API.
    """
    def __init__(self, base_mapping, custom_mapping=None):
        super().__init__(base_mapping)
        self.base_mapping = base_mapping
        self.custom_mapping = custom_mapping or {'exact': {}, 'patterns': []}

    def get(self, key, default=None):
        # Step 1: Check Excel mapping with original key
        if key in self.base_mapping:
            return self.base_mapping[key]
        
        # Step 2: Check custom mapping (exact and wildcard patterns)
        custom_result = apply_custom_mapping(key, self.custom_mapping)
        if custom_result != key:
            return custom_result
        
        # Step 3: Return original key if no mapping found
        return default if default is not None else key

    def __getitem__(self, key):
        result = self.get(key)
        if result is None:
            raise KeyError(key)
        return result


@st.cache_data(show_spinner=False)
def read_ba_mapping_from_excel_bytes(excel_bytes: bytes) -> dict:
    try:
        df = pd.read_excel(BytesIO(excel_bytes))
    except Exception:
        return {}
    cols = {c: c.strip() for c in df.columns}
    df = df.rename(columns=cols)
    ba_cols = [c for c in df.columns if c.lower().startswith("ba")]
    rb_cols = [c for c in df.columns if c.lower().startswith("rb")]
    mapping = {}
    if not ba_cols:
        return mapping
    for _, r in df.iterrows():
        ba_val = str(r.get(ba_cols[0], "")).strip()
        if not ba_val or ba_val.lower() in ["nan", "none"]:
            continue
        for rc in rb_cols:
            rv = r.get(rc)
            if pd.isna(rv):
                continue
            rv_str = str(rv).strip()
            if not rv_str:
                continue
            mapping[rv_str] = ba_val
    return mapping

@st.cache_data(show_spinner=False)
def load_ba_part_names(mapping_path):
    """
    Load BA part names from the mapping file.
    
    Args:
        mapping_path: Path to the BA vs RB mapping Excel file
    
    Returns:
        dict: {rb_part: ba_part_name}
    """
    if not mapping_path.exists():
        return {}
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(mapping_path, read_only=True)
        ws = wb.active
        
        # Get header row
        header_row = [cell.value for cell in ws[1]]
        
        # Find BA partname column and all RB part columns
        ba_name_col_idx = None
        rb_col_indices = []
        
        for idx, col_name in enumerate(header_row):
            if col_name and col_name.strip().lower() == "ba partname":
                ba_name_col_idx = idx
            elif col_name and col_name.strip().lower().startswith("rb part_"):
                rb_col_indices.append(idx)
        
        if ba_name_col_idx is None or not rb_col_indices:
            wb.close()
            return {}
        
        # Build the mapping
        rb_to_ba_name = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            ba_name = str(row[ba_name_col_idx]).strip() if row[ba_name_col_idx] else None
            if not ba_name or ba_name.lower() in ['none', 'nan', 'n/a']:
                continue
            
            for rb_idx in rb_col_indices:
                rb_part = str(row[rb_idx]).strip() if row[rb_idx] else None
                if rb_part and rb_part.lower() not in ['none', 'nan', 'n/a', '']:
                    rb_to_ba_name[rb_part] = ba_name
        
        wb.close()
        return rb_to_ba_name
    
    except Exception as e:
        return {}

@st.cache_data(show_spinner=False)
def load_ba_mapping(mapping_path, custom_mapping_path=None):
    """
    Load BA mapping from Excel file and apply custom rules.

    The mapping process follows this priority:
    1. Check Excel file for explicit mappings (with original RB part)
    2. Check custom mapping CSV (exact and wildcard patterns)
    3. Return original part number if no mapping found

    Parts not found in either Excel or custom mapping will return the original
    RB part number, allowing the application to attempt retrieval from Rebrickable API.

    Args:
        mapping_path: Path to the Excel mapping file
        custom_mapping_path: Path to the custom mapping CSV file (optional)

    Returns:
        EnhancedMapping: Mapping from RB part numbers to BA part numbers
    """
    # Load explicit mappings from Excel
    excel_mapping = {}
    if mapping_path.exists():
        with open(mapping_path, "rb") as f:
            excel_mapping = read_ba_mapping_from_excel_bytes(f.read())

    # Load custom mappings from CSV
    custom_mapping = {'exact': {}, 'patterns': []}
    if custom_mapping_path and custom_mapping_path.exists():
        try:
            custom_df = load_custom_mapping_csv(custom_mapping_path)
            custom_mapping = build_custom_mapping_dict(custom_df)
        except Exception as e:
            st.warning(f"Could not load custom mapping: {e}")

    # Return enhanced mapping that includes custom rules
    return EnhancedMapping(excel_mapping, custom_mapping)



@st.cache_data(show_spinner=False)
def count_parts_in_mapping(mapping_path_str: str, collection_parts: tuple = None, count_type: str = "labels"):
    """
    Count parts in the mapping file.
    
    Args:
        mapping_path_str: String path to mapping file (for caching)
        collection_parts: Tuple of RB part numbers from collection (tuple for hashability)
        count_type: "labels" or "images"
    
    Returns:
        tuple: (total_count, collection_count)
    """
    import openpyxl
    
    mapping_path = Path(mapping_path_str)
    wb = openpyxl.load_workbook(mapping_path)
    ws = wb.active
    header_row = [cell.value for cell in ws[1]]
    
    # Find relevant columns
    if count_type == "labels":
        target_col = header_row.index("BA label URL") + 1 if "BA label URL" in header_row else None
    else:  # images
        target_col = header_row.index("BA partnum") + 1 if "BA partnum" in header_row else None
    
    # Find all RB part columns
    rb_part_cols = []
    for idx, col_name in enumerate(header_row):
        if col_name and col_name.startswith("RB part_"):
            rb_part_cols.append(idx + 1)
    
    total_count = 0
    collection_count = 0
    collection_parts_set = set(collection_parts) if collection_parts else set()
    
    if target_col:
        for row in ws.iter_rows(min_row=2):
            target_val = row[target_col - 1].value
            
            # Check if row is valid
            if count_type == "labels":
                is_valid = target_val and "No label available" not in str(target_val)
            else:  # images
                is_valid = target_val is not None
            
            if is_valid:
                total_count += 1
                
                # Check if any RB part matches collection
                if rb_part_cols and collection_parts_set:
                    for col_idx in rb_part_cols:
                        rb_partnum = str(row[col_idx - 1].value).strip() if row[col_idx - 1].value else None
                        if rb_partnum and rb_partnum in collection_parts_set:
                            collection_count += 1
                            break
    
    wb.close()
    return (total_count, collection_count)


@st.cache_data(show_spinner=False)
def build_ba_to_rb_mapping(mapping_path):
    """
    Build a reverse mapping from BA part numbers to lists of RB part numbers.
    This allows finding similar parts that share the same BA part number.

    Args:
        mapping_path: Path to the BA vs RB mapping Excel file

    Returns:
        dict: {ba_part: [rb_part1, rb_part2, ...]}
    """
    if not mapping_path.exists():
        return {}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(mapping_path, read_only=True)
        ws = wb.active

        # Get header row
        header_row = [cell.value for cell in ws[1]]

        # Find BA partnum column and all RB part columns
        ba_col_idx = None
        rb_col_indices = []

        for idx, col_name in enumerate(header_row):
            if col_name and col_name.strip().lower() == "ba partnum":
                ba_col_idx = idx
            elif col_name and col_name.strip().lower().startswith("rb part_"):
                rb_col_indices.append(idx)

        if ba_col_idx is None or not rb_col_indices:
            wb.close()
            return {}

        # Build the mapping
        ba_to_rb = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            ba_part = str(row[ba_col_idx]).strip() if row[ba_col_idx] else None
            if not ba_part or ba_part.lower() in ['none', 'nan', 'n/a']:
                continue

            rb_parts = []
            for rb_idx in rb_col_indices:
                rb_part = str(row[rb_idx]).strip() if row[rb_idx] else None
                if rb_part and rb_part.lower() not in ['none', 'nan', 'n/a', '']:
                    rb_parts.append(rb_part)

            if rb_parts:
                ba_to_rb[ba_part] = rb_parts

        wb.close()
        return ba_to_rb

    except Exception as e:
        return {}


@st.cache_data(show_spinner=False)
def build_rb_to_similar_parts_mapping(mapping_path):
    """
    Build a mapping from each RB part to all other RB parts that share the same BA part number.
    This is used to find similar/replacement parts.

    Args:
        mapping_path: Path to the BA vs RB mapping Excel file

    Returns:
        dict: {rb_part: [similar_rb_part1, similar_rb_part2, ...]}
    """
    ba_to_rb = build_ba_to_rb_mapping(mapping_path)

    # Build reverse mapping: each RB part maps to all other RB parts with same BA number
    rb_to_similar = {}
    for ba_part, rb_parts in ba_to_rb.items():
        for rb_part in rb_parts:
            # Similar parts are all RB parts for this BA number, excluding itself
            similar = [p for p in rb_parts if p != rb_part]
            if similar:
                rb_to_similar[rb_part] = similar

    return rb_to_similar

